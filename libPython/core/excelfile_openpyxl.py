#!/usr/bin/python
# -*- coding: utf-8 -*-
from __future__ import print_function
if __name__ == '__main__': import _libpythonpath_ #type: ignore (Add libPython\.. into PYTHONPATH when unittest )

import os
import string

import libPython.core.util as util
import openpyxl


class ExcelfileOpenpyxl(object):
    def __init__(self):
        self.filename = None
        self.workbook = None
        self.chartobj_lookup = {}

    """
    -----------------------------------------------
    Open/Save
    -----------------------------------------------
    """
    
    def Open(self, filename, data_only=False, read_only=False):
        if os.path.isfile(filename):
            self.workbook = openpyxl.load_workbook( filename = filename, data_only=data_only, read_only=read_only)
        else:
            self.workbook = openpyxl.Workbook()
            #remove 1st sheet that auto created by "not write_oly open"
            # self.workbook.remove(self.workbook[self.workbook.sheetnames[0]])
        self.filename = filename
    
    def Save(self):
        return self.SaveAs(self.filename)

    def SaveAs(self, out_filename):
        ret = self.workbook.save( filename = out_filename)
        msg = 'save file \"%s\"' % out_filename
        util.LogInfo(msg)
        return ret
        
    """
    -----------------------------------------------
    Properties
    -----------------------------------------------
    """
    def SheetList(self):
        return self.workbook.get_sheet_names()
             
    def GetSheetByName(self, sheet_name):
        try:
            return self.workbook[sheet_name]
        except:
            return None
     
    def GetSheetByIndex(self, index):
        try:
            return self.workbook[self.book.get_sheet_names()[index]]
        except:
            return None

    #...................................
    # max_row = current max valid row #
    # NOTE: max_row is 1 instead of 0 in new created empty sheet
    #...................................
    def max_row(self, sheet_obj=None):
        if sheet_obj:
            return sheet_obj.max_row
        else:
            return self.workbook.active.max_row

    def max_column(self, sheet_obj=None):
        if sheet_obj:
            return sheet_obj.max_row
        else:
            return self.workbook.active.max_column

    #...................................
    # excel column name handling.
    # excel is 1-index (instead of 0-index)
    #...................................
    def col_name(self, col_num):
        retstr=""
        while col_num:
            remainder = (col_num-1) % 26
            retstr = chr(ord("A") + remainder ) + retstr
            col_num = int((col_num-remainder) / 26)
        return retstr

    def col_num(self, col_name):
        num = 0
        for c in col_name:
            if c in string.ascii_letters:
                num = num * 26 + (ord(c.upper()) - ord('A')) + 1
        return num

    """
    -----------------------------------------------
    Operation
    -----------------------------------------------
    """
    def SetActiveSheet(self, sheet_name):
        try:
            index = self.SheetList().index(sheet_name)
            self.workbook.active = index
        except:
            # Create a new sheet
            self.workbook.create_sheet(title=sheet_name)
            index = self.SheetList().index(sheet_name)
            self.workbook.active = index


    """
    -----------------------------------------------
    Write by ROWS
    -----------------------------------------------
    """
    def AppendRows(self, rows ):
        for row in rows:
            self.workbook.active.append(row)

    """
    -----------------------------------------------
    Write by COLS
    -----------------------------------------------
    """
    def AppendCols(self, cols ):
        col_num = len(cols)
        start_col = self.max_column() + 1
        ws = self.workbook.active

        for i, ci in enumerate(range(start_col, start_col+col_num)):
            for j in range(len(cols[i])):
                ws.cell(row=j+1, column=ci).value = cols[i][j]



    """
    ========================================================
    Chart Operation APIs
    ========================================================
    """
    """
    -----------------------------------------------
    Return openpyxl chart object of a designated sheet
    Hiarch:
        {
            sheet_name:{
                chart_type:{
                    group_name:{
                        chart_obj: <openpyxl obj>
                    }
                }
            }
        }
    Naming:
        chart_type: "scatter", "bar"
        chart title: {chart_type}:{group_name}
    -----------------------------------------------
    """
    def GetChartObject(self, sheet_name, chart_type, group_name):
        self.chartobj_lookup[sheet_name] = self.chartobj_lookup.get(sheet_name, {})
        self.chartobj_lookup[sheet_name][chart_type] = self.chartobj_lookup[sheet_name].get(chart_type, {})
        self.chartobj_lookup[sheet_name][chart_type][group_name] = self.chartobj_lookup[sheet_name][chart_type].get(group_name, {})
        chart_info = self.chartobj_lookup[sheet_name][chart_type][group_name]

        if chart_info:
            return chart_info["chart_obj"] #return chart object

        #Calculate chart count
        chart_sum = 0
        for type in self.chartobj_lookup[sheet_name]:
            chart_sum += len(self.chartobj_lookup[sheet_name][type])

        #Create a new chart object
        if chart_type == "scatter":
            chart_info["chart_obj"] = openpyxl.chart.ScatterChart()
            
        elif chart_type == "bar":
            chart_info["chart_obj"] = openpyxl.chart.BarChart()
        else:
            util.LogError(f"Unsupport chart type: {chart_type}")
            return None

        if group_name:
            title = f"{chart_type}:{group_name}"
        else:
            title = f"{chart_type}"
        chart_info["chart_obj"].title = title
        chart_info["id"] = chart_sum - 1
        chart_obj = chart_info["chart_obj"]

        #initial position
        width = 10
        height = 18
        chart_pos = f"{self.col_name( (chart_info['id']%4) * width + 1)}{ int(chart_info['id']/4)*height + 1 }"

        util.LogInfo(f'Create chart object: title={title} id={chart_info["id"]} pos={chart_pos}')

        #Get sheet
        sheet_obj = self.GetSheetByName(sheet_name)
        if not sheet_obj:
            sheet_obj = self.workbook.create_sheet(title=sheet_name)

        #Add chart to sheet (can only do once)
        sheet_obj.add_chart(chart_obj, chart_pos)

        return chart_obj


    """
    -----------------------------------------------
    val_desc_obj = {"sheet":sheet_name, "min_row":2, "min_col":3, "max_col": 65535}
    name_cell = "'primary'!$A$1"
    -----------------------------------------------
    """
    """
    -----------------------------------------------
    Scatter Chart
        StrRefWorkaround:
            1.True : to workaround always set title_from_data=true but actually no
                1-1: in y_value_reference, "min_row" = "min_row"-1
                1-2: only work in vertical arrange data, if horizontal arrange, then "min_col" = "min_col"-1
                1-2: if title_from_data not set to True, can set title to reference string "=Sheet!$A$1"

    -----------------------------------------------
    """
    def AddScatterChart(self, x_val_desc_obj, y_val_desc_obj, name_cell, sheet_name="CHART_GRAPH", group_name="", StrRefWorkaround=True):
        chart_obj = self.GetChartObject(sheet_name, "scatter", group_name)
        series = self.CreateSeries(x_val_desc_obj, y_val_desc_obj, name_cell, StrRefWorkaround=StrRefWorkaround)
        chart_obj.series.append(series)


    def CreateSeries(self, x_val_desc_obj, y_val_desc_obj, name_cell, StrRefWorkaround=True):
        ws = self.GetSheetByName(x_val_desc_obj["sheet"])
        x_val_desc_obj = dict(x_val_desc_obj)
        del x_val_desc_obj["sheet"]
        xvalues = openpyxl.chart.Reference(ws, **x_val_desc_obj)

        ws = self.GetSheetByName(y_val_desc_obj["sheet"])
        y_val_desc_obj = dict(y_val_desc_obj)
        del y_val_desc_obj["sheet"]
        #...................................
        # title_from_data=True workaround
        #...................................
        if StrRefWorkaround:
            y_val_desc_obj["min_row"] = y_val_desc_obj["min_row"]-1
        yvalues = openpyxl.chart.Reference(ws, **y_val_desc_obj)
        series = openpyxl.chart.Series(yvalues, xvalues, title_from_data=True) # NOTE: title_from_data must be set to True for latter StrRef setting
        series.tx.strRef = openpyxl.chart.data_source.StrRef( name_cell )

        return series

    """
    -----------------------------------------------
    val_desc_obj = {"sheet":sheet_name, "min_row":2, "min_col":3, "max_col": 65535}
    name_cell = "'primary'!$A$1"
    -----------------------------------------------
    """
    """
    -----------------------------------------------
    Bar Chart (category (x_value) can only be column )
        from_rows : 
            1. default is False, let data can only be vertical arranged
            2. set to False, horizontal arrange is not good for column count has limit to about 10K while row count is 1000K
        y_include_title:
            1. if include title, auto adjust range(one col/row) to cover title cell
            2. this keep input always define data range
    -----------------------------------------------
    """
    def AddBarChart(self, x_val_desc_obj, y_val_desc_obj, name_cell, sheet_name="CHART_GRAPH", group_name="", from_rows=False, y_include_title=True):
        chart_obj = self.GetChartObject(sheet_name, "bar", group_name)

        # Get Categories
        ws = self.GetSheetByName(x_val_desc_obj["sheet"])
        x_val_desc_obj = dict(x_val_desc_obj)
        del x_val_desc_obj["sheet"]
        xvalues = openpyxl.chart.Reference(ws, **x_val_desc_obj)

        # Get data
        ws = self.GetSheetByName(y_val_desc_obj["sheet"])
        y_val_desc_obj = dict(y_val_desc_obj)
        del y_val_desc_obj["sheet"]

        # if include_title , auto adjust range(one col/row) to cover title cell
        if y_include_title:
            if from_rows:
                y_val_desc_obj["min_col"] = y_val_desc_obj["min_col"]-1
            else:
                y_val_desc_obj["min_row"] = y_val_desc_obj["min_row"]-1

        yvalues = openpyxl.chart.Reference(ws, **y_val_desc_obj)

        # Add to BarChart
        chart_obj.add_data(yvalues, titles_from_data=y_include_title, from_rows=from_rows)
        chart_obj.set_categories(xvalues)
        


        
"""
-----------------------------------------------
Unitest Functions
-----------------------------------------------
"""        
def TEST_1_colname():
    excel = ExcelfileOpenpyxl()
    print(f'{excel.col_num("IVC")}')

def TEST_2_change_active_sheet():
    excel = ExcelfileOpenpyxl()
    # excel.Open("test.xlsx",read_only=True)
    excel.Open("test.xlsx")
    excel.SetActiveSheet("Sheet1")
    print(f"max_row Sheet1={excel.max_row()}")
    excel.AppendRows([["sheet1", "sheet1", "sheet1",]])
    print(f"max_row Sheet1={excel.max_row()}")

    excel.SetActiveSheet("SheetYYYY")
    print(f"max_row SheetY={excel.max_row()}")
    excel.AppendRows([["SheetYYYY", "SheetYYYY", "SheetYYYY",]])
    print(f"max_row SheetY={excel.max_row()}")

    excel.SetActiveSheet("Sheet1")
    print(f"max_row Sheet1={excel.max_row()}")
    excel.AppendRows([["sheet11", "sheet11", "sheet11",]])
    print(f"max_row Sheet1={excel.max_row()}")

    excel.SetActiveSheet("SheetYYYY")
    print(f"max_row SheetY={excel.max_row()}")
    excel.AppendRows([["SheetYYYY2", "SheetYYYY2", "SheetYYYY2",]])
    print(f"max_row SheetY={excel.max_row()}")
    excel.Save()

def TEST_3_chart():
    excel = ExcelfileOpenpyxl()
    excel.Open("test.xlsx")

    for i in range(3):
        x_val_desc_obj = {"sheet":"Sheet", "min_col":1, "min_row":3, "max_row": 7}
        y_val_desc_obj = {"sheet":"Sheet", "min_col":2, "min_row":3, "max_row": 7}
        name_cell = "'sheet'!$A$1"
        excel.AddScatterChart(x_val_desc_obj, y_val_desc_obj, name_cell, group_name=str(i))


        x_val_desc_obj = {"sheet":"Sheet", "min_col":1, "min_row":3, "max_row": 7}
        y_val_desc_obj = {"sheet":"Sheet", "min_col":2, "min_row":3, "max_row": 7}
        name_cell = "'sheet'!$A$1"
        excel.AddBarChart(x_val_desc_obj, y_val_desc_obj, name_cell, group_name=str(i))


    excel.Save()

def TEST_4_AppendCol():
    excel = ExcelfileOpenpyxl()
    excel.Open("test.xlsx")

    cols = [list(range(5))]*3

    print(f"max_col={excel.max_column()}")
    excel.AppendCols(cols)

    excel.Save()






if __name__ == '__main__':
    # TEST_1_colname()
    # TEST_2_change_active_sheet()
    TEST_3_chart()
    # TEST_4_AppendCol()


    
    
    print("Done!")
    
    
    
    pass
