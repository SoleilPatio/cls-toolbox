from openpyxl import load_workbook

class WorkSheet(object):
    def __init__(self, name, ws):
        '''
        :param name: work sheet name
        :param ws: work sheet object from openpyxl
        '''
        self.name = name
        self.work_shet = ws
        
    def GetLookupTable(self):
        '''
        :return ret_table:  0.assume that 1st row in excel is field name 
                            1.one dict object in list for each row
                            2.one dict object represent a row record
        :rtype ret_table: list[dict{"field":value,...}]
        
        '''
        ret_table = []
        ws = self.work_shet
        row = 1 #regard row 1 as field name
        fields = []
        
        for col in range(1, ws.max_column+1):
            content = ws.cell(row=row, column=col).value #.encode(sys.stdin.encoding, "replace").decode(sys.stdin.encoding)
            fields.append(content)
        
        for row in range(2, ws.max_row+1):
            r = {}
            for col in range(1, ws.max_column+1):
                content = ws.cell(row=row, column=col).value #.encode(sys.stdin.encoding, "replace").decode(sys.stdin.encoding)
                
                if col == 1: #regards as comment if leading with "#"
                    if (isinstance(content, unicode) or isinstance(content, str)):
                        if content.startswith('#'):
                            break
                        
                r[fields[col-1]] = content
            
            if len(r):
                ret_table.append(r)    
                    
        return ret_table
        

class ExcelFile(object):
    def __init__(self):
        '''
        :var workbook: workbook object from openpyxl
        :var sheet_names: list["sheet names"]
        :var work_sheets: list[WorkSheet]
        '''
        
        self.workbook = None
        self.sheet_names = None
        self.work_sheets = None

    def Load(self, filename ):
        self.workbook = load_workbook( filename, data_only=True, read_only=True)
        self.sheet_names = self.workbook.get_sheet_names()
        
        self.work_sheets = []
        for wsname in self.sheet_names:
            self.work_sheets.append(WorkSheet(wsname, self.workbook[wsname]))
        
        return self
    
    def GetWorkSheetNames(self):
        return self.sheet_names
            
    def GetWorkSheet(self, name):
        print self.work_sheets
        for ws in self.work_sheets:
            if ws.name == name:
                return ws
        return None
        
        
        
if __name__ == '__main__':
    xlsfile = ExcelFile()
#     xlsfile.Load(r"N:\project\my-mtk-code\ate-script-gen\ate-script-gen-config.xlsx")
    xlsfile.Load(r"N:\project\my-mtk-code\regression-engine\resource\bit_field_definition.xlsx")
    
    ws = xlsfile.GetWorkSheet("Sheet1")
    
    print ws.name
    
    table = ws.GetLookupTable()
    
    for t in table:
        print t
#         print "id=",t["id"], t
        
   
    
    pass