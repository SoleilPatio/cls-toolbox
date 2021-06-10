#!/usr/bin/python
# -*- coding: utf-8 -*-
from __future__ import print_function

import os
import openpyxl
import string
import util


class ExcelfileOpenpyxl(object):
    def __init__(self):
        self.filename = None
        self.workbook = None
        self.current_sheet = None
        pass

    """
    -----------------------------------------------
    Open/Save
    -----------------------------------------------
    """
    
    def Open(self, filename, data_only=False, read_only=False):
        if os.path.isfile(filename):
            self.workbook = openpyxl.load_workbook( filename = filename, data_only=data_only, read_only=read_only)
        else:
            self.workbook = openpyxl.Workbook()
            #remove 1st sheet that auto created by "not write_oly open"
            # self.workbook.remove(self.workbook[self.workbook.sheetnames[0]])
        self.filename = filename
    
    def Save(self):
        return self.SaveAs(self.filename)

    def SaveAs(self, out_filename):
        ret = self.workbook.save( filename = out_filename)
        msg = 'save file \"%s\"' % out_filename
        util.LogInfo(msg)
        return ret
        
    """
    -----------------------------------------------
    Properties
    -----------------------------------------------
    """
    def SheetList(self):
        return self.workbook.get_sheet_names()
             
    def GetSheetByName(self, sheet_name):
        return self.workbook[sheet_name]
     
    def GetSheetByIndex(self, index):
        return self.workbook[self.book.get_sheet_names()[index]]
     
    def max_row(self, sheet_obj=None):
        if sheet_obj:
            return sheet_obj.max_row
        else:
            return self.current_sheet.max_row

    def max_column(self, sheet_obj=None):
        if sheet_obj:
            return sheet_obj.max_row
        else:
            return self.current_sheet.max_column

    #...................................
    # excel column name handling.
    # excel is 1-index (instead of 0-index)
    #...................................
    def col_name(self, col_num):
        retstr=""
        while col_num:
            remainder = (col_num-1) % 26
            retstr = chr(ord("A") + remainder ) + retstr
            col_num = int((col_num-remainder) / 26)
        return retstr

    def col_num(self, col_name):
        num = 0
        for c in col_name:
            if c in string.ascii_letters:
                num = num * 26 + (ord(c.upper()) - ord('A')) + 1
        return num


    """
    -----------------------------------------------
    Write by ROWS
    -----------------------------------------------
    """
    def WriteRows(self, sheet_name, rows ):
        if sheet_name in self.SheetList():
            self.current_sheet = self.GetSheetByName(sheet_name)
        else:
            # Create a new sheet
            self.current_sheet = self.workbook.create_sheet(title=sheet_name)
        self.AppendRows(rows)
    
    def AppendRows(self, rows ):
        for row in rows:
            self.current_sheet.append(row)

    

        
        
        


if __name__ == '__main__':
    excel = ExcelfileOpenpyxl()
    # excel.Open("test.xlsx",read_only=True)
    excel.Open("test.xlsx")

    row_meta = [[123]]
    excel.WriteRows("meta",row_meta)
    row_pri = [[321, "=A1+'meta'!$A$1"]]
    excel.WriteRows("pri",row_pri)
    excel.Save()
    
    print("Done!")
    
    
    
    pass