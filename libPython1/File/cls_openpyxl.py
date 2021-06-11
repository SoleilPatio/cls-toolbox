from openpyxl import load_workbook
from clsobject import clsobject


class clsOpenpyxlFile(clsobject):
    def __init__(self):
        self.book = None
        pass
    
    def Open(self, filename):
        self.book = load_workbook( filename, data_only=True, read_only=True)
        
    def ListOpenExcelFiles(self):
        print "N/A!"
                 
    def ListSheets(self):
        print self.book.get_sheet_names()
        return self.book.get_sheet_names()
             
    def GetSheetByName(self, sheet_name):
        return self.book[sheet_name]
     
    def GetSheetByIndex(self, index):
        return self.book[self.book.get_sheet_names()[index]]
     
    def LastRowOfSheet(self, sheet_obj):
        return sheet_obj.max_row

    """
    cascade_field_name: if field_name_row_num is > 0, want to cascade filed name from 0-field_name_row_num ?
    pass_down_field_names: if current cell do not have value, use value in upper cell
    """
    def ParseSheetAsTable(self, sheet_obj, field_name_row_num = 0, cascade_field_name = True, pass_down_field_names = [], lower_case_field_name = False):
        ret_table = []
        
        
                
        """
        decide field_names
        """
        if cascade_field_name:
            start_row_num = 0
        else:
            start_row_num = field_name_row_num
             
        field_names = ["" for _ in range(sheet_obj.max_column) ]
        for i in range(start_row_num, field_name_row_num+1 ):
            last_value = ""
            row_values = [cell.value for cell in sheet_obj[i+1]] #note that openpyxl is 1-indexed
            for j, value in enumerate(row_values): 
                if value == "" or value is None:
                    value = last_value
                if field_names[j] != "" and field_names[j] is not None:
                    field_names[j] = field_names[j] + "%"

                if lower_case_field_name:
                    cur_str = str(value).lower()
                else:
                    cur_str = str(value)
                
                field_names[j] = field_names[j]+ cur_str
                last_value = value
                
        
        
        """
        decide pass down field index
        """
        pass_down_index = []
        for i, name in enumerate(field_names):
            for pdfn in pass_down_field_names:
                if pdfn.lower() == name.lower():
                    pass_down_index.append(i)
        
        print "cascade_field_name=", cascade_field_name
        print "field_names=",field_names      
        print "pass_down_index=", pass_down_index    
            
            
        
        max_row = self.LastRowOfSheet(sheet_obj)
        max_col = len(field_names)
#         print "rowcount=", max_row
#         print "colcount=", max_col
        
        
        """
        Read content
        """
        #last pass-down-field-value
        last_pdfv = {}
        for i in pass_down_index:
            last_pdfv[i] = ""
        
        for r in range(field_name_row_num+1, max_row):
            r_dict = {}
            row_values = [cell.value for cell in sheet_obj[r+1]] #note that openpyxl is 1-indexed
            for i, value in enumerate(row_values):
                if i in pass_down_index:
                    if value == "" or value is None:
                        value = last_pdfv[i]
                    else:
                        last_pdfv[i] = value
                r_dict[field_names[i]] = value
            ret_table.append(r_dict)
             
        return ret_table
            
        
        
        
        


if __name__ == '__main__':
    filename = r"D:\Project-D\TASKs\coda_parser.task\MY_NEW_TEST_Rocky_dip_a_000_ctl.xlsx"
    
    openpyxlobj = clsOpenpyxlFile()
    
    openpyxlobj.Open(filename)
    
    openpyxlobj.ListOpenExcelFiles()
    openpyxlobj.ListSheets()
    sheet = openpyxlobj.GetSheetByIndex(4)
    print "last row of ", openpyxlobj.ListSheets()[4], " is ",  openpyxlobj.LastRowOfSheet(sheet)
     
    table = openpyxlobj.ParseSheetAsTable(sheet, field_name_row_num = 1 , pass_down_field_names = ["register%name"])
      
    for r in table:
        print [ "%s = %s"%( x.split("%")[-1], r[x]) for x in r if r[x] != "" ]
 
    target = [ x for x in table if x["Field%power element"] ]
     
    for x in target:
        print x
     
    
    
    print "Done!"
    
    
    
    pass